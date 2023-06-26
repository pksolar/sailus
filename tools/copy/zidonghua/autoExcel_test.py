import pandas as pd
from openpyxl import load_workbook
"""
r"#E:\code\python_PK\bleeding\img_process\17_R1C78_resize1.3\res\Lane01\Lane01_fastq.fq_total.out"
AveReadsLength	100.0
TotalReads	299279
MappedReads	272515	0.9106	1.0
UniqMappedReads	264013	0.8822	0.9688
EffectiveReads	141403	0.4725	0.5189
ConcordantReads	145963	0.4877	0.5356
MismatchReads	126293	0.422	0.4634

8行数据。


"""
import pandas as pd
from openpyxl import load_workbook


# 打开 Excel 文件
book = load_workbook('output_333.xlsx')
data =
# 选择要操作的工作表
writer = pd.ExcelWriter('output_333.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
sheet_name = 'Sheet1'

# 将数据添加到指定位置
worksheet = writer.sheets[sheet_name]
print(worksheet.max_row)

# 保存并关闭 Excel 文件
writer.save()