import pandas as pd

from openpyxl import load_workbook
import openpyxl
import xlsxwriter

# 打开文本文件

name_xlsx = "output_total_result"
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Sheet1"
wb.save(f'{name_xlsx}.xlsx')


machineNames = ["30_resize","1.9.1_resize","1.9.39_resize","22_resize","08_resize","17_R1C78_resize"]#"1.9.1_resize","1.9.39_resize",

dict_resize = {       "_ori":[2160,4096],
                         1.15:[2484,4710],
                      1.2:[2592,4914],
                       1.25:[2700,5120],
                       1.3:[2808,4324],
                       1.35:[2916,5530],
                       1.4:[3024,5734]   }

hang = 0
for machine_name in machineNames:
    for key, value in dict_resize.items():
        data_list = []
        try:
            path = rf'E:\data\resize_test\{machine_name}{key}\res\Lane01\Lane01_fastq.fq_total.out'
            name = path.split("\\")[-4]
            with open(path, 'r') as file:
                # 读取文件内容并拆分为行
                lines = file.readlines()[:15]
            # 提取逗号分隔的值
            for i,line in enumerate (lines):
                fields = line.strip().split('\t')
                #print(fields)
                if i>1 and i<8:
                    data_list = data_list + fields[1:]
                    #print(data_list)
                if i == 14:
                    data_list = data_list+fields[1:]

            #data_list里有16个数。第一个数在 3B,  最后一个数在 3Q
            data_list_number = []
            for data in data_list:
                data = float(data)
                if data < 1 :
                    data = data * 100
                data_list_number.append(round(data,3))

            # 将文本读入 Pandas 数据框
            # 打开 Excel 文件
            book = load_workbook(f'{name_xlsx}.xlsx')

            # 选择要操作的工作表
            writer = pd.ExcelWriter(f'{name_xlsx}.xlsx', engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            sheet_name = 'Sheet1'

            # 将数据添加到指定位置
            worksheet = writer.sheets[sheet_name]
            max_row  = worksheet.max_row
            print(max_row)
            for i, value in enumerate(data_list_number):
                worksheet.cell(row=max_row+1, column=i, value=value)
                worksheet.cell(row=max_row+1,column=1 ,value=name)
            hang += 1

            writer.save()
        except:
            print(f"error: {machine_name},{key}")
